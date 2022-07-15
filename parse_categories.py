#!/usr/bin/env python
# -*- coding: utf-8 -*-

import csv
import json
import re
import sys
from openpyxl import load_workbook

WORKBOOK_FILENAME = "Output_Category_Mapping_2021.xlsx"
SELECTED_TABLES_CSV = "selected-tables.csv"

def expand_range(code, sep):
    endpoints = code.split(sep)
    return [i for i in range(int(endpoints[0]), int(endpoints[1]) + 1)]

def with_ranges_expanded(codes):
    result = []
    for code in codes:
        if code == '':
            continue
        elif code.startswith('-'):
            if '-' in code[1:]:
                raise ValueError("Unexpected range beginning with a negative number: " + code)
            result.append(code)
        elif '-' in code:
            result.extend(expand_range(code, '-'))
        elif '>' in code:
            result.extend(expand_range(code, '>'))
        elif '–' in code:
            result.extend(expand_range(code, '–'))
        else:
            if re.compile('^[0-9]*$').match(code) is None:
                raise ValueError("Unexpected code: " + code)
            result.append(int(code))
    return result

def parse_category_codes(codes_str):
    codes = str(codes_str).replace(" ", "").strip().split(",")
    codes = with_ranges_expanded(codes)
    return codes

def extract_categories(sheet, column):
    extra_strings = []
    categories = []

    found_categories = False
    row = 2
    while row < 1000:
        val = sheet.cell(row, column).value
        if re.search('^[-0-9]', str(val)):
            codes_str = str(val).strip().replace(" ", "")
            codes = parse_category_codes(val)
            categories.append({
                "codes": codes,
                "name": sheet.cell(row, column + 1).value
            })
        elif val is not None:
            extra_strings.append(str(val))
        elif found_categories:
            break
        row += 1
    return {"extra_strings": extra_strings, "categories": categories}

def parse_sheet(sheet, var_details_map, selected_vars):
    sheet_title = sheet.title.upper()
    if sheet_title not in var_details_map:
        return None

    title_as_in_index = var_details_map[sheet_title]["2021 Mnemonic (variable)"]
    if title_as_in_index not in selected_vars:
        return None

    selected_categories = selected_vars[title_as_in_index]
    number_of_selected_categories_parsed = 0    # to check we found all of the requested cats

    rows = [row for row in sheet.rows]

    classifications = []
    selected_default = False

    for cell in rows[0]:
        if isinstance(cell.value, str) and re.search('[A-Z]', cell.value):
            short_category_code = cell.value.strip().split("_")[-1]
            if selected_categories[0] == "all" or short_category_code in selected_categories:
                classification = extract_categories(sheet, cell.column)
                if not selected_default and len(classification["categories"]) < 12:
                    selected_default = True
                    classification["default"] = True
                else:
                    classification["default"] = False
                classification["classification_code"] = cell.value
                classifications.append(classification)
                number_of_selected_categories_parsed += 1

    if selected_categories[0] != "all" and number_of_selected_categories_parsed != len(selected_categories):
        print("Unexpected number of categories", title_as_in_index, sheet_title)
        raise 1

    return {"metadata": var_details_map[sheet_title], "classifications": classifications}


def parse_index_table(sheet):
    var_details_map = {}
    column_names = []

    col = 1
    while True:
        value = sheet.cell(2, col).value
        if value is None:
            break
        column_names.append(value)
        col += 1

    row = 3
    while sheet.cell(row, 1).value is not None:
        for i, column_name in enumerate(column_names):
            column = i + 1
            value = sheet.cell(row, column).value
            if value is None:
                value = ""
            if column == 1:
                value = value.upper()
                sheet_name = value
                if sheet.cell(row, column).hyperlink is not None:
                    sheet_name = sheet.cell(row, column).hyperlink.location.split('!')[0].upper().replace("'", "")
                d = {}
                var_details_map[sheet_name] = d
            d[column_name] = value
        row += 1

    return var_details_map

def parse_selected_vars(selected_tables_csv):
    selected_vars = {}
    with open(selected_tables_csv, "r") as f:
        csvreader = csv.reader(f)
        for i, (classifications, mnemonic) in enumerate(csvreader):
            if i == 0:
                continue
            classifications = classifications.split(",")
            selected_vars[mnemonic.upper()] = classifications
    return selected_vars

def main():
    workbook_filename = sys.argv[1] or WORKBOOK_FILENAME
    selected_tables_csv = sys.argv[2] or SELECTED_TABLES_CSV

    selected_vars = parse_selected_vars(selected_tables_csv)

    wb = load_workbook(workbook_filename)
    sheetnames = wb.sheetnames

    var_details_map = parse_index_table(wb['INDEX'])

    variable_sheetnames = sheetnames[3:]
    variables = []

    for sheetname in variable_sheetnames:
        var_data = parse_sheet(wb[sheetname], var_details_map, selected_vars)
        if var_data is not None:
            variables.append(var_data)

    print(json.dumps(variables, indent=4))
    #print(len(variables))
    #print(sum(len(v["classifications"]) for v in variables))

    #for v in variables:
    #    print(len(v["classifications"]))

if __name__ == "__main__":
    main()
